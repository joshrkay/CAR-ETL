"""
OpenPyXL Parser - Understanding Plane

Parser implementation for Excel spreadsheets using openpyxl.
Fallback parser for .xlsx files when pandas fails.
"""
import logging
from io import BytesIO

from src.exceptions import ParserError

from .base import BaseParser, ExtractedTable, ParseResult

logger = logging.getLogger(__name__)


class OpenPyXLParser(BaseParser):
    """Parser implementation using openpyxl for Excel files."""

    def __init__(self) -> None:
        """Initialize OpenPyXL parser."""
        pass

    async def parse(self, content: bytes, mime_type: str) -> ParseResult:
        """
        Parse Excel spreadsheet using openpyxl.

        Args:
            content: Raw document bytes
            mime_type: MIME type of the document

        Returns:
            ParseResult with extracted content

        Raises:
            ParserError: If parsing fails
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ParserError("openpyxl", "openpyxl library not installed")

        excel_file = BytesIO(content)

        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception as e:
            logger.exception("Failed to load Excel workbook")
            raise ParserError("openpyxl", f"Failed to load Excel file: {str(e)}")

        text_parts = []
        tables = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract all cell values
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                rows_data.append(list(row))

            if rows_data:
                # First row as headers
                headers = [str(cell) if cell is not None else "" for cell in rows_data[0]]

                # Remaining rows as data
                rows = []
                for row_data in rows_data[1:]:
                    row_dict = {
                        headers[i] if i < len(headers) else f"col_{i}": str(cell) if cell is not None else ""
                        for i, cell in enumerate(row_data)
                    }
                    rows.append(row_dict)

                # Build text representation
                text_parts.append(
                    f"Sheet: {sheet_name}\n" +
                    "\n".join(["\t".join(map(str, row)) for row in rows_data])
                )

                tables.append(ExtractedTable(
                    table_name=sheet_name,
                    headers=headers,
                    rows=rows,
                    page_number=None,
                    confidence=1.0
                ))

        return ParseResult(
            text="\n\n".join(text_parts),
            pages=[],
            tables=tables,
            metadata={"sheet_count": len(workbook.sheetnames)},
            parser_confidence=1.0
        )

    async def health_check(self) -> bool:
        """
        Check if openpyxl is available.

        Returns:
            True if openpyxl is installed, False otherwise
        """
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False
